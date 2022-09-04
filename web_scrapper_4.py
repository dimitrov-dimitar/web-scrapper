import multiprocessing
import re
import subprocess
from datetime import date, datetime

import requests
from bs4 import BeautifulSoup
from decouple import Config, RepositoryEnv
from openpyxl import load_workbook

import price_alert
from hotels import info

DOTENV_FILE = ".env.dev"
# DOTENV_FILE = "/home/dimitar/Desktop/web_scrapper/.env.prod"

config = Config(RepositoryEnv(DOTENV_FILE))


def find_between(all_content, first, last):
    try:
        start = all_content.index(first) + len(first)
        end = all_content.index(last, start)
        return all_content[start:end]
    except ValueError:
        return ""


# Hotel Scrapping

workbook = load_workbook(config.get("WORKBOOK"))
sheet = workbook.active

max_row = sheet.max_row
max_col = sheet.max_column

today = date.today()
day = today.strftime("%d.%m.%Y")

now = datetime.now()
current_time = now.strftime("%H:%M:%S")

def hotel_info_scraping():
    for name, value in info.items():

        h_url = value[0]
        h_filter = value[1]
        h_sheet = value[2]

        result = requests.get(h_url).text
        doc = BeautifulSoup(result, "html.parser")

        # 23.09.2022, 7 days
        # rest_day = "17.06.2022"
        rest_day = "23.09.2022"

        information = doc.find_all(
            "a", {"title": re.compile(rf"Направи запитване за {rest_day}")}
        )
        # print(information)
        information = str(information)
        hotel_str = find_between(information, h_filter, " €")

        hotel_c = h_sheet
        hotel_r = max_row + 1

        # Append day
        sheet.cell(column=hotel_c, row=hotel_r, value=day)

        # Append time
        sheet.cell(column=hotel_c + 1, row=hotel_r, value=current_time)

        # Skip empty fields
        if not hotel_str:
            continue
        price = int(hotel_str[-4:])
        sheet.cell(column=hotel_c + 2, row=hotel_r, value=price)

        # print(f"{hotel} = {price} €")

    workbook.save(config.get("WORKBOOK"))
    workbook.close()

    print(day, current_time)
    print("Web Scrapping Successful")
    print(f"Information for rest day - {rest_day}")

hotel_info_scraping()
# p1 =  multiprocessing.Process(target= hotel_info_scraping)
# p2 =  multiprocessing.Process(target= hotel_info_scraping)
# p3 =  multiprocessing.Process(target= hotel_info_scraping)
# p4 =  multiprocessing.Process(target= hotel_info_scraping)
# p1.start()
# p2.start()
# p3.start()
# p4.start()
# p1.join()
# p2.join()
# p3.join()
# p4.join()

price_alert.price()
subprocess.call(config.get("UPLOAD_SCRIPT"))
subprocess.call('/home/dimitar/Desktop/web_scrapper/ifttt.sh')

print("----------------------------------------------------------------------")
