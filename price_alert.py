from decouple import Config, RepositoryEnv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import mail
from hotels import info
from movements import down, up


DOTENV_FILE = ".env.dev"
# DOTENV_FILE = "/home/dimitar/Desktop/web_scrapper/.env.prod"

config = Config(RepositoryEnv(DOTENV_FILE))


def price():
    workbook = load_workbook(config.get("WORKBOOK"))
    sheet = workbook.active

    max_row = sheet.max_row
    max_col = sheet.max_column
    total_msg = ""
    subject = ""
    hotel_counter = 0

    for name, value in info.items():
        h_sheet = value[2]

        hotel_c = h_sheet
        hotel_r = max_row

        current_price = sheet.cell(
            column=hotel_c + 2,
            row=hotel_r,
        ).value
        old_price = sheet.cell(column=hotel_c + 2, row=hotel_r - 1).value

        # Skip empty fields
        if not current_price or not old_price:
            continue
        diff_price = int(current_price) - int(old_price)

        sign = ""
        current_cell = sheet.cell(
            column=hotel_c + 2,
            row=hotel_r,
        )
        if diff_price > 0:
            sign = "+"
            current_cell.fill = PatternFill(
                start_color="00ff00", fill_type="solid"
            )
            subject += up
        elif diff_price < 0:
            current_cell.fill = PatternFill(
                start_color="ff0000", fill_type="solid"
            )
            subject += down
        else:
            sign = ""

        if diff_price != 0:
            hotel_counter += 1
            msg = (
                f"{hotel_counter}. {name} - Current Price: {current_price}€, Old Price: {old_price}€\n"
                f"Price difference: {sign}{diff_price}€\n"
            )
            # print(msg)
            total_msg += msg

    workbook.save(config.get("WORKBOOK"))
    workbook.close()

    if total_msg:
        print(f"Price Alert Detected")
        print(total_msg)
        subject += f" {config.get('ENV')}: Turkey, Antalya Price Alert"
        mail.send(total_msg, subject)
    else:
        print("No price change")
